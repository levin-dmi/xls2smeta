from rowparser import RowParser
from openpyxl import load_workbook
from openpyxl.styles import Font
import collections
import configparser
import re
from glob import glob


def col_name2num(name: str) -> int:
    """
    Translate Excel column name to number
    """
    num = ord(name[0]) - ord('A') + 1
    if len(name) > 1:
        num = num * 26 + ord(name[1]) - ord('A') + 1
    return num


# Открываем файл конфигурации и читаем ее
task = configparser.ConfigParser(inline_comment_prefixes=('#',))
task.read('xls2smeta.conf')
xls_in_columns = {'name': col_name2num(task['global']['name']),
                  'brand': col_name2num(task['global']['brand']),
                  'code': col_name2num(task['global']['code']),
                  'producer': col_name2num(task['global']['producer']),
                  'unit': col_name2num(task['global']['unit']),
                  'quantity': col_name2num(task['global']['quantity']), }
res = col_name2num(task['global']['result'])
xls_out_columns = {'system': res,
                   'name': res + 1,
                   'result': res + 2,
                   'brand': res + 3,
                   'code': res + 4,
                   'producer': res + 5,
                   'unit': res + 6,
                   'quantity': res + 7,
                   'comment': res + 8, }


recognized_row = 0
all_row = 0
normalized_words = {}
for filename in glob(task['global']['file']):
    wb = load_workbook(filename, read_only=False, data_only=True)
    ws = wb.active

    # Обрабатываем строки с объединенными ячейками, содержимое в верхней
    merged_cells = ws.merged_cells.ranges.copy()
    for cell_range in merged_cells:
        for column in xls_in_columns.values():
            top_v = ws.cell(row=cell_range.bounds[1], column=column).value
            for r in range(cell_range.bounds[1]+1, cell_range.bounds[3]+1):
                v = ws.cell(row=r, column=column).value
                if v:
                    if top_v:
                        top_v = str(top_v) + str(v)
                    else:
                        top_v = v
                ws.cell(row=r, column=column, value=None)
            ws.cell(row=cell_range.bounds[1], column=column, value=top_v)
        ws.unmerge_cells(start_row=cell_range.bounds[1], start_column=cell_range.bounds[0],
                         end_row=cell_range.bounds[3], end_column=cell_range.bounds[2])

    detector = RowParser()
    system_name = '?'

    i = 0
    for row in ws.rows:
        # Читаем нужные ячейки строки в словарь
        data: dict = {'name': row[xls_in_columns['name'] - 1].value,
                      'brand': row[xls_in_columns['brand'] - 1].value,
                      'code': row[xls_in_columns['code'] - 1].value,
                      'producer': row[xls_in_columns['producer'] - 1].value,
                      'unit': row[xls_in_columns['unit'] - 1].value,
                      'quantity': row[xls_in_columns['quantity'] - 1].value, }
        if data['quantity']:
            all_row += 1

        # пробуем распознать
        res_list = detector.parse(data)
        for res_row in res_list:
            i += res_row['incoming_rows']  # Пропускаем несколько строчек если нам вернули 1 строку на неколько

            if res_row['type'] == detector.DETECTED_RT:  # считаем детекнутые строки
                recognized_row += 1
                bold = True
            else:
                bold = False
                if res_row['type'] == detector.NORMALIZED_RT and len(res_row['material'].split()) > 0:
                    nw = res_row['material'].split()[0]
                    if nw in normalized_words:
                        normalized_words[nw] += 1
                    else:
                        normalized_words[nw] = 1

            ws.cell(row=i, column=xls_out_columns['system'], value=res_row['system']).font = Font(bold=bold)
            ws.cell(row=i, column=xls_out_columns['name'], value=res_row['material']).font = Font(bold=bold)
            ws.cell(row=i, column=xls_out_columns['result'], value=res_row['dimension']).font = Font(bold=bold)
            ws.cell(row=i, column=xls_out_columns['brand'], value=res_row['brand']).font = Font(bold=bold)
            ws.cell(row=i, column=xls_out_columns['code'], value=res_row['code']).font = Font(bold=bold)
            ws.cell(row=i, column=xls_out_columns['producer'], value=res_row['producer']).font = Font(bold=bold)
            ws.cell(row=i, column=xls_out_columns['unit'], value=res_row['unit']).font = Font(bold=bold)
            if res_row['quantity']:
                ws.cell(row=i, column=xls_out_columns['quantity'], value=res_row['quantity']).font = Font(bold=bold)
            # ws.cell(row=i, column=xls_out_columns['comment'], value=res_row['type']).font = Font(bold=bold)

    # Добиваем последнюю строчку
    data: dict = {'name': '', 'brand': '', 'code': '', 'producer': '', 'unit': '', 'quantity': '', }
    res_list = detector.parse(data)
    for res_row in res_list:
        i += res_row['incoming_rows']  # Пропускаем несколько строчек если нам вернули 1 строку на неколько

        if res_row['type'] == detector.DETECTED_RT:
            recognized_row += 1
            bold = True
        else:
            bold = False

        ws.cell(row=i, column=xls_out_columns['system'], value=res_row['system']).font = Font(bold=bold)
        ws.cell(row=i, column=xls_out_columns['name'], value=res_row['material']).font = Font(bold=bold)
        ws.cell(row=i, column=xls_out_columns['result'], value=res_row['dimension']).font = Font(bold=bold)
        ws.cell(row=i, column=xls_out_columns['brand'], value=res_row['brand']).font = Font(bold=bold)
        ws.cell(row=i, column=xls_out_columns['code'], value=res_row['code']).font = Font(bold=bold)
        ws.cell(row=i, column=xls_out_columns['producer'], value=res_row['producer']).font = Font(bold=bold)
        ws.cell(row=i, column=xls_out_columns['unit'], value=res_row['unit']).font = Font(bold=bold)
        if res_row['quantity']:
            ws.cell(row=i, column=xls_out_columns['quantity'], value=res_row['quantity']).font = Font(bold=bold)
        # ws.cell(row=i, column=xls_out_columns['comment'], value=res_row['type']).font = Font(bold=bold)

    wb.save('./result/' + re.sub(r"./data/", "", filename))
print('Распознано позиций:', recognized_row, 'из', all_row)
c = collections.Counter(normalized_words)
print(c.most_common(len(c)*2//10))


# TODO Сделать условие - какие листы обрабатываем из книги: первый, все, с конкретным именем
# TODO Везде добавить исключения и проверки на корректность данных в конфиге
# TODO При проблемах с объединенными ячейками
#  см https://config9.com/apps/excel/how-to-detect-merged-cells-in-excel-with-openpyxl/
# TODO Можно удалять столбцы через VBA но только под Windows https://habr.com/ru/post/232291/#com
# TODO Нужно добавить os чтобы все работало и в винде
# TODO можно переделать детектор систем, чтобы он искал все варианты систем в строке и автоматически делал название
# TODO попробовать Pandas для ускорения работы с Excel
# TODO https://blog.davep.org/2018/06/02/a_little_speed_issue_with_openpyxl.html
# TODO xlwings как быстрая альтернатива?

